import openpyxl


def countrows(file, sheetname):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheetname]
    rows = sh.max_row
    return rows


def countcolumns(file, sheetname):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheetname]
    cols = sh.max_column
    return cols


def readData(file, sheetname, rows, columns):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheetname]
    data = sh.cell(rows, columns)
    d = data.value
    return d


def writeData(file, sheetname, row, column, value):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheetname]
    cell = sh.cell(row=row, column=column)
    cell.value = value
    wb.save(file)
    return cell.value



